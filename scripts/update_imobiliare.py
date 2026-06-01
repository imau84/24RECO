#!/usr/bin/env python3
"""
Actualizare automată imobiliare_data.json cu date ANCPI.
Rulează săptămânal (Luni) via GitHub Actions.
Detectează lunile noi de pe ancpi.ro și adaugă datele în JSON.
"""

import json
import re
import sys
import time
import urllib.request
import io
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

# ─── Config ──────────────────────────────────────────────────────────────────

DATA_FILE = Path("public/imobiliare_data.json")

# Mapare luni RO → număr
MONTHS_MAP = {
    "ianuarie": "01", "februarie": "02", "martie": "03", "aprilie": "04",
    "mai": "05", "iunie": "06", "iulie": "07", "august": "08",
    "septembrie": "09", "octombrie": "10", "noiembrie": "11", "decembrie": "12",
}

# Județe în ordinea din Excel ANCPI (42 rânduri de date)
JUDETE_ORDER = [
    "ALBA", "ARAD", "ARGEȘ", "BACAU", "BIHOR", "BISTRIȚA NĂSĂUD", "BOTOȘANI",
    "BRĂILA", "BRAȘOV", "BUCUREȘTI", "BUZĂU", "CĂLĂRAȘI", "CARAȘ SEVERIN",
    "CLUJ", "CONSTANȚA", "COVASNA", "DAMBOVIȚA", "DOLJ", "GALAȚI", "GIURGIU",
    "GORJ", "HARGHITA", "HUNEDOARA", "IALOMIȚA", "IASI", "ILFOV", "MARAMUREȘ",
    "MEHEDINȚI", "MUREȘ", "NEAMȚ", "OLT", "PRAHOVA", "SĂLAJ", "SATU MARE",
    "SIBIU", "SUCEAVA", "TELEORMAN", "TIMIȘ", "TULCEA", "VÂLCEA", "VASLUI",
    "VRANCEA",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.ancpi.ro/",
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def fetch_url(url: str, retries: int = 3) -> bytes:
    for i in range(retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as r:
                return r.read()
        except Exception as e:
            if i < retries - 1:
                print(f"  Retry {i+1}/{retries}: {e}")
                time.sleep(3 * (i + 1))
            else:
                raise


def normalize_judet(name: str) -> str:
    """Normalizează numele județului din Excel la cheia din JSON."""
    name = str(name).strip().upper()
    # Fix variante comune
    replacements = {
        "BISTRIȚA-NĂSĂUD": "BISTRIȚA NĂSĂUD",
        "BISTRITA NASAUD": "BISTRIȚA NĂSĂUD",
        "CARAȘ-SEVERIN": "CARAȘ SEVERIN",
        "DÂMBOVIȚA": "DAMBOVIȚA",
        "IAȘI": "IASI",
        "MUREȘ": "MUREȘ",
        "VÂLCEA": "VÂLCEA",
    }
    return replacements.get(name, name)


def parse_int(val) -> int:
    try:
        return int(str(val).strip().replace(" ", "").replace(",", "") or 0)
    except (ValueError, TypeError):
        return 0


# ─── Excel parsers ───────────────────────────────────────────────────────────

def parse_vanzari(xlsx_bytes: bytes) -> dict:
    """
    Parsează Excel vânzări județ.
    Col: Nr.crt | Județ | Ext.Agr | Ext.Neagr | Intr.Constr | Intr.Fara | Unit.Indiv | Total
    Date încep de la rândul 6 (index 5, 0-based).
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    result = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        # Rândul de date are nr. curent (int) în col 0
        try:
            nr = int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue
        if nr < 1 or nr > 42:
            continue

        judet = normalize_judet(row[1] or "")
        if not judet:
            continue

        result[judet] = {
            "extrav_agr":    parse_int(row[2]),
            "extrav_neagr":  parse_int(row[3]),
            "intrav_constr": parse_int(row[4]),
            "intrav_fara":   parse_int(row[5]),
            "unitati_indiv": parse_int(row[6]),
            "total":         parse_int(row[7]),
        }

    print(f"  Vânzări: {len(result)} județe parsate")
    return result


def parse_resedinta(xlsx_bytes: bytes) -> dict:
    """
    Parsează Excel reședință județ.
    Col: Nr.crt | Județ | UAT | Ext.Agr | Ext.Neagr | Intr.Constr | Intr.Fara | Unit.Indiv | Total
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active
    result = {}

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            nr = int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue
        if nr < 1 or nr > 42:
            continue

        judet = normalize_judet(row[1] or "")
        uat = str(row[2] or "").strip()
        if not judet:
            continue

        result[judet] = {
            "uat":           uat,
            "extrav_agr":    parse_int(row[3]),
            "extrav_neagr":  parse_int(row[4]),
            "intrav_constr": parse_int(row[5]),
            "intrav_fara":   parse_int(row[6]),
            "unitati_indiv": parse_int(row[7]),
            "total":         parse_int(row[8]),
        }

    print(f"  Reședință: {len(result)} județe parsate")
    return result


# ─── ANCPI page scraper ───────────────────────────────────────────────────────

def get_download_ids(page_url: str) -> dict:
    """
    Extrage ID-urile de download din pagina ANCPI.
    Returnează {'vanzari': id, 'resedinta': id} sau None dacă pagina nu există.
    """
    try:
        html = fetch_url(page_url).decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"  Pagina nu există sau eroare: {e}")
        return None

    # Caută linkuri de download
    pattern = r'download\.php\?id=(\d+)[^"]*"[^>]*>([^<]+)</a'
    matches = re.findall(pattern, html)

    ids = {}
    for dl_id, text in matches:
        text_lower = text.lower().strip()
        if "vanzari" in text_lower and "resedinta" not in text_lower and "vanzari" not in ids:
            ids["vanzari"] = dl_id
        elif "resedinta" in text_lower and "resedinta" not in ids:
            ids["resedinta"] = dl_id

    if not ids:
        # Fallback: ia primele 2 linkuri de download în ordine
        all_ids = re.findall(r'download\.php\?id=(\d+)', html)
        unique = list(dict.fromkeys(all_ids))  # deduplicate păstrând ordinea
        if len(unique) >= 2:
            ids["vanzari"] = unique[0]
            ids["resedinta"] = unique[1]

    return ids if len(ids) == 2 else None


def build_page_url(year: int, month_num: str) -> tuple[str, str]:
    """Construiește URL-ul paginii de statistică ANCPI pentru o lună dată."""
    month_name_ro = {
        "01": "ianuarie", "02": "februarie", "03": "martie", "04": "aprilie",
        "05": "mai", "06": "iunie", "07": "iulie", "08": "august",
        "09": "septembrie", "10": "octombrie", "11": "noiembrie", "12": "decembrie",
    }
    name = month_name_ro[month_num]
    url = f"https://www.ancpi.ro/statistica-{name}-{year}/"
    key = f"{year}_{month_num}"
    return url, key


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=== Actualizare imobiliare_data.json ===\n")

    # Citește JSON-ul curent
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    existing_raw = set(data.get("RAW", {}).keys())
    existing_res = set(data.get("RES", {}).keys())
    print(f"Luni existente RAW: {sorted(existing_raw)}")
    print(f"Luni existente RES: {sorted(existing_res)}\n")

    # Determină lunile de verificat: ultimele 3 luni + următoarele 2
    import datetime
    today = datetime.date.today()
    months_to_check = []
    for delta in range(-2, 3):  # 3 luni în urmă → 2 luni în viitor
        d = today.replace(day=1)
        # Calculează luna relativă
        total_months = d.year * 12 + d.month - 1 + delta
        y = total_months // 12
        m = total_months % 12 + 1
        months_to_check.append((y, str(m).zfill(2)))

    added = []
    for year, month_num in months_to_check:
        key = f"{year}_{month_num}"

        # Verifică dacă avem deja ambele seturi
        has_raw = key in existing_raw
        has_res = key in existing_res
        if has_raw and has_res:
            print(f"[{key}] Deja complet, skip.")
            continue

        page_url, _ = build_page_url(year, month_num)
        print(f"[{key}] Verificare {page_url}")

        ids = get_download_ids(page_url)
        if not ids:
            print(f"[{key}] Pagina nu există sau fără linkuri, skip.\n")
            continue

        print(f"  IDs: vanzari={ids['vanzari']}, resedinta={ids['resedinta']}")

        # Descarcă și parsează vânzări
        if not has_raw:
            try:
                dl_url = f"https://www.ancpi.ro/wp-content/plugins/download-attachments/includes/download.php?id={ids['vanzari']}"
                print(f"  Descarcă vânzări: {dl_url}")
                xlsx = fetch_url(dl_url)
                raw_data = parse_vanzari(xlsx)
                if raw_data:
                    data.setdefault("RAW", {})[key] = raw_data
                    added.append(f"{key} RAW")
            except Exception as e:
                print(f"  EROARE vânzări: {e}")

        # Descarcă și parsează reședință
        if not has_res:
            try:
                dl_url = f"https://www.ancpi.ro/wp-content/plugins/download-attachments/includes/download.php?id={ids['resedinta']}"
                print(f"  Descarcă reședință: {dl_url}")
                xlsx = fetch_url(dl_url)
                res_data = parse_resedinta(xlsx)
                if res_data:
                    data.setdefault("RES", {})[key] = res_data
                    added.append(f"{key} RES")
            except Exception as e:
                print(f"  EROARE reședință: {e}")

        print()

    if added:
        # Sortează cheile în JSON
        data["RAW"] = dict(sorted(data["RAW"].items()))
        data["RES"] = dict(sorted(data["RES"].items()))
        if "SECT" in data:
            data["SECT"] = dict(sorted(data["SECT"].items()))

        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

        print(f"✅ Adăugate: {', '.join(added)}")
        print(f"   Fișier salvat: {DATA_FILE}")
    else:
        print("ℹ️  Nicio dată nouă găsită.")

    return len(added)


if __name__ == "__main__":
    n = main()
    sys.exit(0 if n >= 0 else 1)
